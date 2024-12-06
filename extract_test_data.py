import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def extract_test_data(folder_path):
    # Regular expression to extract Test Case IDs
    tc_pattern = re.compile(r'self\.th\.protocol\.start\(\s*["\']TC-(AT-)?(\d+)["\']')
    end_pattern = re.compile(r'self\.th\.protocol\.end\(\s*\)')

    # List to store the extracted data
    results = []

    # Get the name of the selected folder for Folder Path column
    folder_name = os.path.basename(folder_path)
    print(f"Processing folder: {folder_name}")

    # Ensure the folder exists
    if not os.path.exists(folder_path):
        print(f"Folder does not exist: {folder_path}")
        return []

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.py'):
                script_name = os.path.splitext(file)[0]
                file_path = os.path.join(root, file)

                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                except Exception as e:
                    print(f"    Error reading file {file}: {e}")
                    continue

                # Initialize variables for parsing
                current_tc = None
                inside_tc = False
                found_matches = False  # Flag to track if any matches are found

                # Debug: Print current file being processed
                print(f"Processing file: {script_name}")

                # Process each line in the file
                for i, line in enumerate(lines):
                    # Match the Test Case ID
                    if tc_match := tc_pattern.search(line):
                        found_matches = True
                        if current_tc:  # Save the previous test case if any
                            results.append({
                                'Folder Path': folder_name,
                                'Script Name': script_name,
                                'Test Case ID': f"TC-{current_tc}"
                            })
                        # Start a new test case
                        current_tc = tc_match.group(2)
                        inside_tc = True
                        print(f"  Found TC: TC-{current_tc} at line {i + 1}")

                    # Match the end of the test case
                    if inside_tc and end_pattern.search(line):
                        found_matches = True
                        if current_tc:  # Save the completed test case
                            results.append({
                                'Folder Path': folder_name,
                                'Script Name': script_name,
                                'Test Case ID': f"TC-{current_tc}"
                            })
                            print(f"  End of TC-{current_tc} at line {i + 1}")
                        # Reset for the next potential test case
                        current_tc = None
                        inside_tc = False

                # If no matches were found, still add the file to the results
                if not found_matches:
                    print(f"  No test cases found in: {script_name}")
                    results.append({
                        'Folder Path': folder_name,
                        'Script Name': script_name,
                        'Test Case ID': None
                    })

    return results


# Folders to process
folder_names = [
    "design_validation",
    "full_market_release",
    "g7",
    "limited_market_release",
    "long_execution"
]

# Base folder path
base_folder_path = "/Users/juliapinedo/Desktop/op5ios-atm/omnipod_5_ios"

# Aggregate all results into a single list
all_results = []
for folder_name in folder_names:
    folder_path = os.path.join(base_folder_path, folder_name)
    folder_results = extract_test_data(folder_path)
    all_results.extend(folder_results)

# Convert results to a DataFrame
if all_results:
    df = pd.DataFrame(all_results, columns=['Folder Path', 'Script Name', 'Test Case ID'])

    # Sort by Script Name (alphabetically), then by Folder Path (custom order)
    df = df.sort_values(by=['Script Name', 'Folder Path'], key=lambda col: col.map({
        "design_validation": 1,
        "full_market_release": 2,
        "g7": 3,
        "limited_market_release": 4,
        "long_execution": 5
    }).fillna(0) if col.name == "Folder Path" else col)

    # Add new columns
    df['Status'] = None
    df['Bug # (if applicable)'] = None

    # Save DataFrame to Excel
    output_file = "all_tcs.xlsx"
    df.to_excel(output_file, index=False)

    # Add data validation for the "Status" column
    wb = load_workbook(output_file)
    ws = wb.active

    # Create data validation for "Status"
    dv = DataValidation(type="list", formula1='"Fail,Pass,Blocked"', allow_blank=True)
    dv.error = "Invalid status selected"
    dv.errorTitle = "Invalid Input"
    dv.prompt = "Select a status"
    dv.promptTitle = "Status Selection"

    # Apply the data validation to the Status column
    status_col = ws['D']
    for cell in status_col[1:]:  # Skip header
        dv.add(cell)
    ws.add_data_validation(dv)

    # Save the workbook
    wb.save(output_file)
    print(f"Data extraction completed. Results saved in: {output_file}")
else:
    print("No test cases found to save.")
